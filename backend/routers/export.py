import csv
import io
import json
from collections import defaultdict
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, Response, Request
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from fpdf import FPDF
from ..api_utils import get_output_dir
from ..audit import record, AuditAction, AuditResource
from ..database import get_db
from .. import models
from ..rls import is_postgres, resolve_user_tenant_id, set_tenant_id
from ..local import settings as local_settings
from ..security.path_safety import sanitize_filename, safe_path
from ..services.export import (
    export_balance_sheet,
    export_general_ledger,
    export_profit_loss,
    export_trial_balance,
    export_transactions,
)
from ..utils.redaction import mask_account_number, mask_transaction_description
from .auth import get_current_user

router = APIRouter(prefix="/export", tags=["export"])

def _resolve_tenant_id(request: Request, current_user: models.User, tenant_id: int | None = None) -> int:
    if tenant_id is not None:
        return tenant_id
    if local_settings.is_single_user():
        return resolve_user_tenant_id(current_user)
    raise HTTPException(status_code=400, detail="tenant_id query parameter is required")


def _wrap_tenant(request: Request, db: Session, current_user: models.User):
    if not is_postgres():
        return
    tenant_id = request.headers.get("x-tenant-id")
    if tenant_id:
        set_tenant_id(db, int(tenant_id))
        return
    set_tenant_id(db, resolve_user_tenant_id(current_user))


def _tenant_id_or_404(tenant_id: int | None, current_user: models.User) -> int:
    return _resolve_tenant_id(None, current_user, tenant_id)

@router.get("/transactions")
def export_transactions_csv(
    request: Request,
    tenant_id: int | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _wrap_tenant(request, db, current_user)
    tid = _tenant_id_or_404(tenant_id, current_user)
    csv_content = export_transactions(db, tid, current_user.id, start_date, end_date)
    record(db, current_user, AuditAction.EXPORT, AuditResource.EXPORT, None,
           {"type": "transactions", "tenant_id": tid, "start_date": start_date, "end_date": end_date})
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=transactions.csv"}
    )

@router.get("/general-ledger")
def export_general_ledger_csv(
    request: Request,
    tenant_id: int | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _wrap_tenant(request, db, current_user)
    tid = _tenant_id_or_404(tenant_id, current_user)
    csv_content = export_general_ledger(db, tid, current_user.id, start_date, end_date)
    record(db, current_user, AuditAction.EXPORT, AuditResource.EXPORT, None,
           {"type": "general_ledger", "tenant_id": tid, "start_date": start_date, "end_date": end_date})
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=general_ledger.csv"}
    )

@router.get("/trial-balance")
def export_trial_balance_csv(
    request: Request,
    tenant_id: int | None = None,
    as_of: str | None = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _wrap_tenant(request, db, current_user)
    tid = _tenant_id_or_404(tenant_id, current_user)
    if not as_of:
        raise HTTPException(status_code=400, detail="as_of query parameter is required")
    csv_content = export_trial_balance(db, tid, current_user.id, as_of)
    record(db, current_user, AuditAction.EXPORT, AuditResource.EXPORT, None,
           {"type": "trial_balance", "tenant_id": tid, "as_of": as_of})
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=trial_balance.csv"}
    )

@router.get("/profit-loss")
def export_profit_loss_csv(
    request: Request,
    tenant_id: int | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _wrap_tenant(request, db, current_user)
    tid = _tenant_id_or_404(tenant_id, current_user)
    if not start_date or not end_date:
        raise HTTPException(status_code=400, detail="start_date and end_date query parameters are required")
    csv_content = export_profit_loss(db, tid, current_user.id, start_date, end_date)
    record(db, current_user, AuditAction.EXPORT, AuditResource.EXPORT, None,
           {"type": "profit_loss", "tenant_id": tid, "start_date": start_date, "end_date": end_date})
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=profit_loss.csv"}
    )

@router.get("/balance-sheet")
def export_balance_sheet_csv(
    request: Request,
    tenant_id: int | None = None,
    as_of: str | None = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _wrap_tenant(request, db, current_user)
    tid = _tenant_id_or_404(tenant_id, current_user)
    if not as_of:
        raise HTTPException(status_code=400, detail="as_of query parameter is required")
    csv_content = export_balance_sheet(db, tid, current_user.id, as_of)
    record(db, current_user, AuditAction.EXPORT, AuditResource.EXPORT, None,
           {"type": "balance_sheet", "tenant_id": tid, "as_of": as_of})
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=balance_sheet.csv"}
    )

@router.get("/formats")
def export_formats():
    return [
        {"id": "csv", "name": "CSV", "description": "Comma-separated values", "icon": "file-spreadsheet", "color": "green", "status": "ready"},
        {"id": "json", "name": "JSON", "description": "JSON array of transactions", "icon": "file-json", "color": "blue", "status": "ready"},
        {"id": "qif", "name": "QIF", "description": "Quicken Interchange Format", "icon": "file-text", "color": "purple", "status": "ready"},
        {"id": "qbo", "name": "QuickBooks Online CSV", "description": "QBO-ready CSV", "icon": "file-spreadsheet", "color": "green", "status": "ready"},
        {"id": "xero", "name": "Xero CSV", "description": "Xero-compatible CSV", "icon": "file-spreadsheet", "color": "blue", "status": "ready"},
        {"id": "excel", "name": "Excel", "description": "Microsoft Excel workbook", "icon": "file-spreadsheet", "color": "green", "status": "ready"},
        {"id": "pdf", "name": "PDF Summary", "description": "PDF reconciliation summary", "icon": "file-text", "color": "red", "status": "ready"},
        {"id": "parquet", "name": "Parquet", "description": "Apache Parquet for data science", "icon": "file-binary", "color": "orange", "status": "ready"},
    ]

def _statement_export_filename(statement_id: int, format: str) -> str:
    """Return a sanitized filename for a statement export."""
    safe_format = sanitize_filename(format)
    return f"statement_{statement_id}.{safe_format}"

@router.get("/statement/{statement_id}")
def export_statement(request: Request,
                     statement_id: int,
                     format: str = "csv",
                     db: Session = Depends(get_db),
                     current_user: models.User = Depends(get_current_user)):
    _wrap_tenant(request, db, current_user)
    statement = db.query(models.Statement).filter(
        models.Statement.id == statement_id,
        models.Statement.user_id == current_user.id
    ).first()
    if not statement:
        raise HTTPException(status_code=404, detail="Statement not found")

    transactions = db.query(models.Transaction).filter(
        models.Transaction.statement_id == statement_id,
        models.Transaction.tenant_id == statement.tenant_id
    ).order_by(models.Transaction.date.asc()).all()

    # --- 1. JSON ---
    if format == "json":
        data = [{
            "id": t.id,
            "date": t.date,
            "description": mask_transaction_description(t.description),
            "amount": float(t.amount), "type": t.tx_type, "category": t.category,
            "running_balance": float(t.running_balance) if t.running_balance is not None else None
        } for t in transactions]
        return Response(
            content=json.dumps(data, indent=2),
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename={_statement_export_filename(statement_id, "json")}"}
        )

    # --- 2. RAW CSV ---
    elif format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["id","date","description","amount","type","category","running_balance"])
        for t in transactions:
            writer.writerow([
                t.id, t.date, mask_transaction_description(t.description),
                float(t.amount), t.tx_type, t.category,
                float(t.running_balance) if t.running_balance is not None else ""
            ])
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={_statement_export_filename(statement_id, "csv")}"}
        )

    # --- 3. QUICKBOOKS ONLINE (QBO) CSV ---
    elif format == "qbo":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Date", "Description", "Withdrawals", "Deposits"])
        for t in transactions:
            try:
                dt = datetime.strptime(t.date, "%Y-%m-%d")
                qbo_date = dt.strftime("%m/%d/%Y")
            except (ValueError, TypeError):
                qbo_date = t.date

            amount = float(t.amount)
            withdrawals = abs(amount) if amount < 0 else ""
            deposits = amount if amount > 0 else ""
            writer.writerow([qbo_date, mask_transaction_description(t.description), withdrawals, deposits])

        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={_statement_export_filename(statement_id, "qbo.csv")}"}
        )

    # --- 4. XERO CSV ---
    elif format == "xero":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Date", "Payee", "Description", "Reference", "Amount"])
        for t in transactions:
            writer.writerow([t.date, mask_transaction_description(t.description), t.category, t.id, float(t.amount)])

        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={_statement_export_filename(statement_id, "xero.csv")}"}
        )

    # --- 5. EXCEL (.xlsx) ---
    elif format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"

        headers = ["Date", "Description", "Category", "Amount", "Type", "Running Balance"]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill

        for t in transactions:
            ws.append([
                t.date, mask_transaction_description(t.description), t.category, float(t.amount),
                t.tx_type, float(t.running_balance) if t.running_balance is not None else None
            ])

        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=4).number_format = '"$"#,##0.00'
            ws.cell(row=row, column=6).number_format = '"$"#,##0.00'

        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value is not None), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={_statement_export_filename(statement_id, "xlsx")}"}
        )

    # --- 6. PDF SUMMARY REPORT ---
    elif format == "pdf":
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 10, "TaxFlow Pro - Statement Summary", new_x="LMARGIN", new_y="NEXT", align="C")
        pdf.ln(10)

        pdf.set_font("Helvetica", "B", 12)
        # FIXED: Use multi_cell for long filenames to prevent overflow
        pdf.multi_cell(0, 8, f"Statement: {statement.filename}")
        pdf.multi_cell(0, 8, f"Period: {statement.period_start or 'N/A'} to {statement.period_end or 'N/A'}")
        pdf.ln(5)

        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, "Reconciliation", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 12)
        open_bal = float(statement.opening_balance) if statement.opening_balance is not None else 0.0
        close_bal = float(statement.closing_balance) if statement.closing_balance is not None else 0.0
        variance = float(statement.variance) if statement.variance is not None else 0.0
        balanced = "Yes" if statement.is_balanced else "No"

        pdf.cell(90, 8, f"Opening Balance: ${open_bal:,.2f}", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(90, 8, f"Closing Balance: ${close_bal:,.2f}", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(90, 8, f"Variance: ${variance:,.2f}", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(90, 8, f"Balanced: {balanced}", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(5)

        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, "Category Summary", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 12)

        cat_totals = defaultdict(float)
        for t in transactions:
            cat_totals[t.category] += float(t.amount)

        for cat, total in sorted(cat_totals.items(), key=lambda x: abs(x[1]), reverse=True):
            pdf.multi_cell(0, 8, f"{cat}: ${total:,.2f}")

        # FIXED: dest='B' returns raw bytes for FastAPI Response
        pdf_bytes = pdf.output(dest='B')
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={_statement_export_filename(statement_id, "summary.pdf")}"}
        )

    # --- 7. PARQUET (Data Science / ML) ---
    elif format == "parquet":
        try:
            import pandas as pd
        except ImportError:
            raise HTTPException(status_code=500, detail="pandas not installed")

        data = [{
            "id": t.id,
            "date": t.date,
            "description": mask_transaction_description(t.description),
            "amount": float(t.amount), "tx_type": t.tx_type, "category": t.category,
            "running_balance": float(t.running_balance) if t.running_balance is not None else None
        } for t in transactions]

        df = pd.DataFrame(data)
        output_parquet = io.BytesIO()
        df.to_parquet(output_parquet, engine="pyarrow", index=False)
        output_parquet.seek(0)

        return Response(
            content=output_parquet.getvalue(),
            media_type="application/octet-stream",
            headers={"Content-Disposition": f"attachment; filename={_statement_export_filename(statement_id, "parquet")}"}
        )

    # --- 8. QIF (Legacy) ---
    elif format == "qif":
        lines = ["!Account", "NExported", "^", "!Type:Bank"]
        for t in transactions:
            lines.append(f"D{t.date}")
            lines.append(f"P{mask_transaction_description(t.description)}")
            lines.append(f"T{t.amount}")
            lines.append(f"L{t.category}")
            lines.append("^")
        return Response(
            content="\n".join(lines),
            media_type="text/plain",
            headers={"Content-Disposition": f"attachment; filename={_statement_export_filename(statement_id, "qif")}"}
        )

    else:
        raise HTTPException(status_code=400, detail="Format must be json, csv, qbo, xero, excel, pdf, parquet, or qif")
